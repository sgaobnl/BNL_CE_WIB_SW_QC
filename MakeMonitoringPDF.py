from wib_cfgs import WIB_CFGS
import time
import sys
import numpy as np
import pickle
import copy
import time, datetime, random, statistics    
import csv
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from fpdf import FPDF

import openpyxl as px
import statsmodels.api as sm

plot_dir= "./output/plots/"

def linear_fit(x, y):
    error_fit = False 
    try:
        results = sm.OLS(y,sm.add_constant(x)).fit()
    except ValueError:
        error_fit = True 
    if ( error_fit == False ):
        error_gain = False 
        try:
            slope = results.params[1]
        except IndexError:
            slope = 0
            error_gain = True
        try:
            constant = results.params[0]
        except IndexError:
            constant = 0
    else:
        slope = 0
        constant = 0
        error_gain = True

    y_fit = np.array(x)*slope + constant
    delta_y = abs(y - y_fit)
    inl = delta_y / (max(y)-min(y))
    peakinl = max(inl)
    return slope, constant, peakinl, error_gain


    

def GetPlotMeasuredLArASICThroughMonitoringPin(FE = 1, BL = 200):
    plt.close()
    inputFile = "./tmp_data/measure_LArASIC_through_monitoring_pin_{}mVBL.bin".format(BL)
    with open(inputFile, 'rb') as f:
        rawData = pickle.load(f)

    tempPointsX = []
    tempPointsY = []
    for i in range(16 * 8):
        if (rawData[i][0] != FE):
            continue
        tempPointsX.append(rawData[i][1]) #channel
        tempPointsY.append(rawData[i][2])

    fig, ax = plt.subplots()
    line1, = ax.plot(tempPointsX, tempPointsY)
    ax.grid(True)
    plt.xlabel('channels')
    plt.ylabel('V')
    ax.set_title('FE' + str(FE) + ', measure LArASIC {}mV BL through Monitoring pin'.format(BL))
    plt.savefig(plot_dir+ "FE{}_LArASIC_".format(FE) + "_monitoring_pin.jpg")
    plt.close()

def GetPlotMeasuredTemperatureThroughMonitoringPin():
    plt.close()
    inputFile = "./tmp_data/measure_temperature_through_monitoring_pin.bin"
    with open(inputFile, 'rb') as f:
        rawData = pickle.load(f)

    tempPointsX = []
    tempPointsY = []
    for fe in range(8):
        tempPointsX.append('FE{}'.format(fe))
        tempPointsY.append(rawData[fe][1])

    fig, ax = plt.subplots()
    line1, = ax.plot(tempPointsX, tempPointsY)
    ax.grid(True)
    plt.xlabel('FEs')
    plt.ylabel('V')
    ax.set_title('measure Temperatue through Monitoring pin')
    plt.savefig(plot_dir+ "temperature_monitoring_pin.jpg")
    plt.close()

def GetPlotMeasuredVGBRThroughMonitoringPin():
    plt.close()
    inputFile = "./tmp_data/measure_VBGR_through_monitoring_pin.bin"
    with open(inputFile, 'rb') as f:
        rawData = pickle.load(f)

    tempPointsX = []
    tempPointsY = []
    for fe in range(8):
        tempPointsX.append('FE{}'.format(fe))
        tempPointsY.append(rawData[fe][1])

    fig, ax = plt.subplots()
    line1, = ax.plot(tempPointsX, tempPointsY)
    ax.grid(True)
    plt.xlabel('FEs')
    plt.ylabel('V')
    ax.set_title('measure VBGR through Monitoring pin')
    plt.savefig(plot_dir+ "VBGR_monitoring_pin.jpg")
    plt.close()

def GetPlotMeasuredVGBRThroughVBGRPin():
    plt.close()
    inputFile = "./tmp_data/measure_VBGR_through_VBGR_pin.bin"
    with open(inputFile, 'rb') as f:
        rawData = pickle.load(f)

    tempPointsX = []
    tempPointsY = []
    for fe in range(8):
        tempPointsX.append('FE{}'.format(fe))
        tempPointsY.append(rawData[fe][1])

    fig, ax = plt.subplots()
    line1, = ax.plot(tempPointsX, tempPointsY)
    ax.grid(True)
    plt.xlabel('FEs')
    plt.ylabel('V')
    ax.set_title('measure VBGR through VBGR pin')
    plt.savefig(plot_dir+ "VBGR_monitoring_pin.jpg")
    plt.close()

def GetPlotMeasuredLArASICDACThroughMonitoringPin(FE = 0):
    plt.close()
    inputFile = ["","","","",""]
    rawData = [[],[],[],[],[]]
    #0: disable gain
    #1: 14mV/fc
    #2: 4.7
    #3: 7.8
    #4: 25mV/fc
    amp_RT = []
    inputFile[0] = "./tmp_data/measure_LArASIC_DAC_through_monitoring_pin_sg0_0_sg1_0_sgp_1.bin"
    inputFile[1] = "./tmp_data/measure_LArASIC_DAC_through_monitoring_pin_sg0_0_sg1_0_sgp_0.bin"
    inputFile[2] = "./tmp_data/measure_LArASIC_DAC_through_monitoring_pin_sg0_1_sg1_0_sgp_0.bin"
    inputFile[3] = "./tmp_data/measure_LArASIC_DAC_through_monitoring_pin_sg0_0_sg1_1_sgp_0.bin"
    inputFile[4] = "./tmp_data/measure_LArASIC_DAC_through_monitoring_pin_sg0_1_sg1_1_sgp_0.bin"
    for i in range(5):
        with open(inputFile[i], 'rb') as f:
            rawData[i] = pickle.load(f)


    for files in range(5):
        tempDAC = []
        for i in range(64 * 8):
            if (rawData[files][i][0] != FE):
                continue
            tempDAC.append(rawData[files][i][2] * 1000)
        amp_RT.append(np.array(tempDAC))

    amp_RT_SGP1=amp_RT[0]
    amp_RT_SGP0_14=amp_RT[1]
    amp_RT_SGP0_47=amp_RT[2]
    amp_RT_SGP0_78=amp_RT[3]
    amp_RT_SGP0_25=amp_RT[4]

    remove_val=0
    dac_lin=64
    encs = [0] * (dac_lin-remove_val) 
    ps = [0] * (dac_lin-remove_val) 
    lin_amp_200mv_ch4_rt_si_47=(dac_lin-remove_val)*[0]
    lin_amp_200mv_ch4_rt_si_78=(dac_lin-remove_val)*[0]
    lin_amp_200mv_ch4_rt_si_14=(dac_lin-remove_val)*[0]
    lin_amp_200mv_ch4_rt_si_25=(dac_lin-remove_val)*[0]
    lin_amp_200mv_ch4_rt_si_sgp1=(dac_lin-remove_val)*[0]

    amp_200mv_ch4=(dac_lin-2)*[0]

    temp=  "RT"#"LN2"#"RT"
    Gain= ""#"25"  #"14" ## #"7.8"
    sgp= "" # "SGP_1"

    a = []
    for i in range(64):
            a.append(i)
    dacbins = np.array(a)

    x = np.array([63,62,61,60,59,58,57,56,55,54,53,52,51,50,49,48,47,46,45,44,43,42,41,40,39,38,37,36,35,34,33,32,31,30,29,28,27,26,25,24,23,22,21,20,19,18,17,16,15,14,13,12,11,10,9,8,7,6,5,4,3,2,1,0])

    plt.plot(dacbins, amp_RT_SGP0_47, marker='.', linewidth=1.5, label= "SGP = 0, 4.7 mV/fC")
    plt.plot(dacbins, amp_RT_SGP0_78, marker='.', linewidth=1.5, label= "SGP = 0, 7.8 mV/fC")
    plt.plot(dacbins, amp_RT_SGP0_14, marker='.', linewidth=1.5, label= "SGP = 0, 14 mV/fC")
    plt.plot(dacbins, amp_RT_SGP0_25, marker='.', linewidth=1.5, label= "SGP = 0, 25 mV/fC")
    plt.plot(dacbins, amp_RT_SGP1, marker='.', linewidth=1.5, label= "SGP = 1")
    plt.legend()

    plt.ylabel("Amplitude / mV")
    plt.xlabel("DAC bins")
    plt.title("FE{} DAC, ".format(FE) +  temp   )
    plt.tight_layout()
    plt.grid()
    plt.savefig(plot_dir+ "FE{}_DAC_".format(FE) + temp + "_" + sgp+ "_" + Gain + "mV-fc.jpg")
    plt.close()

    remove_val=2
    amp=amp_RT_SGP0_47
    dnl1 = []
    for i in range(len(amp) - remove_val -1 ):
        dnl1.append( amp[i+1+remove_val] -amp[i+remove_val] )
    print (dnl1)
    lsb = ((amp[63-remove_val]-amp[0] )/ (63-remove_val))
    dnl1 = ( ( dnl1 - lsb) /lsb ) 
    print(np.mean(dnl1))

    remove_val=0
    amp=amp_RT_SGP0_78
    dnl2 = []
    for i in range(len(amp) - remove_val -1 ):
        dnl2.append( amp[i+1+remove_val] -amp[i+remove_val] )
    print (dnl2)
    lsb =  ((amp[63-remove_val]-amp[0] )/ (63-remove_val))
    dnl2 = ( ( dnl2 - lsb) /lsb )
    print(np.mean(dnl2))

    remove_val=0
    amp=amp_RT_SGP0_14
    dnl3 = []
    for i in range(len(amp) - remove_val -1 ):
        dnl3.append( amp[i+1+remove_val] -amp[i+remove_val] )
    print (dnl3)
    lsb =  ((amp[63-remove_val]-amp[0] )/ (63-remove_val))
    dnl3 = ( ( dnl3 - lsb) /lsb )
    print(np.mean(dnl3))

    remove_val=0
    amp=amp_RT_SGP0_25
    dnl4 = []
    for i in range(len(amp) - remove_val -1 ):
        dnl4.append( amp[i+1+remove_val] -amp[i+remove_val] )
    print (dnl4)
    lsb =  ((amp[63-remove_val]-amp[0] )/ (63-remove_val))
    dnl4 = ( ( dnl4 - lsb) /lsb )
    print(np.mean(dnl4))
    remove_val=2
    amp=amp_RT_SGP1
    dnl5 = []
    for i in range(len(amp) - remove_val -1 ):
        dnl5.append( amp[i+1+remove_val] -amp[i+remove_val] )
    print (dnl5)
    lsb =  ((amp[63-remove_val]-amp[0] )/ (63-remove_val))
    dnl5 = ( ( dnl5 - lsb) /lsb )
    print(np.mean(dnl5))

    plt.plot(dacbins[(1+2):], dnl1[:],  marker='.', linewidth=1.5, label= "SGP = 0, 4.7 mV/fC")
    plt.plot(dacbins[(1+0):], dnl2,  marker='.', linewidth=1.5, label= "SGP = 0, 7.8 mV/fC")
    plt.plot(dacbins[(1+0):], dnl3,  marker='.', linewidth=1.5, label= "SGP = 0, 14 mV/fC")
    plt.plot(dacbins[(1+0):], dnl4,  marker='.', linewidth=1.5, label= "SGP = 0, 25 mV/fC")
    plt.plot(dacbins[(1+2):], dnl5[:],  marker='.', linewidth=1.5, label= "SGP = 1")
    plt.legend()
    plt.ylabel("DNL / LSB")
    plt.xlabel("DAC bins")
    plt.ylim([-1,1])
    plt.text(min(x)+remove_val, 0.6, "DNL(i) = (V(i) - V(i-1) - LSB) / LSB" )
    plt.title("FE{} DAC DNL, ".format(FE) +  temp + " " + sgp + " " + Gain + "" )
    plt.grid(True)
    plt.savefig(plot_dir+ "FE{}_DNL_".format(FE) + temp + "_" + sgp+ "_" + Gain + "mV-fc.jpg")
    plt.close()

    inl1 = []
    for i in range(len(dnl1)):
        if (i == 0):
            inl1.append(0)
        else:
            inl1.append(np.sum(dnl1[0:i]))
    print (inl1)

    inl2 = []
    for i in range(len(dnl2)):
        if (i == 0):
            inl2.append(0)
        else:
            inl2.append(np.sum(dnl2[0:i]))
    print (inl2)

    inl3 = []
    for i in range(len(dnl3)):
        if (i == 0):
            inl3.append(0)
        else:
            inl3.append(np.sum(dnl3[0:i]))
    print (inl3)

    inl4 = []
    for i in range(len(dnl4)):
        if (i == 0):
            inl4.append(0)
        else:
            inl4.append(np.sum(dnl4[0:i]))
    print (inl4)

    inl5 = []
    for i in range(len(dnl5)):
        if (i == 0):
            inl5.append(0)
        else:
            inl5.append(np.sum(dnl5[0:i]))
    print (inl5)

    plt.plot(dacbins[(1+2):], inl1[:], marker='.', linewidth=1.5, label= "SGP = 0, 4.7 mV/fC")
    plt.plot(dacbins[(1+0):], inl2,  marker='.', linewidth=1.5, label= "SGP = 0, 7.8 mV/fC")
    plt.plot(dacbins[(1+0):], inl3,  marker='.', linewidth=1.5, label= "SGP = 0, 14 mV/fC")
    plt.plot(dacbins[(1+0):], inl4,  marker='.', linewidth=1.5, label= "SGP = 0, 25 mV/fC")
    plt.plot(dacbins[(1+2):], inl5[:],  marker='.', linewidth=1.5, label= "SGP = 1")
    plt.legend()

    plt.ylabel("INL / LSB")
    plt.xlabel("DAC bins")
    plt.ylim([-3,3])
    plt.text(min(x)+ remove_val, 2.1, "INL(i) = DNL(0) + DNL(1) + ... + DNL(i-1)" )
    plt.title("FE{} DAC INL, ".format(FE) +  temp + " " + sgp + " " + Gain + "" )
    plt.grid(True)
    plt.savefig(plot_dir+ "FE{}_INL_".format(FE) + temp + "_" + sgp+ "_" + Gain + "mV-fc.jpg")
    plt.close()


#GetPlotMeasuredLArASICThroughMonitoringPin(FE = 0, BL = 200)
#GetPlotMeasuredLArASICThroughMonitoringPin(FE = 2, BL = 900)
GetPlotMeasuredTemperatureThroughMonitoringPin()
GetPlotMeasuredVGBRThroughMonitoringPin()
GetPlotMeasuredVGBRThroughVBGRPin()
for i in range(8):
    GetPlotMeasuredLArASICThroughMonitoringPin(FE = i, BL = 200)
    GetPlotMeasuredLArASICThroughMonitoringPin(FE = i, BL = 900)
    GetPlotMeasuredLArASICDACThroughMonitoringPin(FE = i)


def MakePDF(feNumber):
    print ('making result pdf for fe{}'.format(feNumber))
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Times', '', 24)
    pdf.cell(0, 20, 'Power Measurement', 0, 1, 'C')
    pdf.set_font('Times', '', 12)
    FE = 'FE' + str(feNumber) + '_'
    FE_VDD = FE + 'VDD'
    FE_VDDO = FE + 'VDDO'
    FE_VPPP = FE + 'VPPP'
    FE_V = [FE_VDD, FE_VDDO, FE_VPPP]
    tempConfigurationIndex = 0
    for d in rawData:
        if (tempConfigurationIndex == 0): 
            pdf.cell(60, 10, 'SE=OFF SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 1): 
            pdf.cell(60, 10, 'SE=ON SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 2): 
            pdf.cell(60, 10, 'SE=OFF SEDC=ON', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 3): 
            pdf.cell(60, 10, 'SE=OFF SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 4): 
            pdf.cell(60, 10, 'SE=ON SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 5): 
            pdf.cell(60, 10, 'SE=OFF SEDC=ON', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        tempConfigurationIndex += 1
        pdf.cell(60, 10, '', 1, 1, 'C')
        pdf.cell(60, 10, 'Power Rail', 1, 0, 'C')
        pdf.cell(60, 10, 'Voltages (V)', 1, 0, 'C')
        pdf.cell(60, 10, 'Current (mA)', 1, 1, 'C')
        tempPower = 0
        tempIndex = 0;

        for v in FE_V:
            pdf.cell(60, 10, v, 1, 0, 'C')
            pdf.cell(60, 10, str(d[v][0]), 1, 0, 'C')
            pdf.cell(60, 10, str(d[v][1]), 1, 1, 'C')
            tempPower += d[v][2]
        pdf.cell(60, 10, 'Power (mW/Ch)', 1, 0, 'C')
        pdf.cell(60, 10, str(tempPower), 1, 0, 'C')
        pdf.cell(60, 10, '', 1, 1, 'C')
        pdf.cell(0, 10, '', 0, 1, 'C')

    tempPlot = GetCurrentPlot(feNumber,0)
    tempPlot.savefig('./output/plots/FE{}_900mV.png'.format(feNumber))
    tempPlot = GetCurrentPlot(feNumber,1)
    tempPlot.savefig('./output/plots/FE{}_200mV.png'.format(feNumber))
    pdf.image('./output/plots/FE{}_200mV.png'.format(feNumber), x = None, y = None, w = 150, h = 100, type = 'png', link = '')
    pdf.image('./output/plots/FE{}_900mV.png'.format(feNumber), x = None, y = None, w = 150, h = 100, type = 'png', link = '')

    pdf.output('./output/result_FE{}.pdf'.format(feNumber), 'F')


def MakePDFAll():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Times', '', 24)
    pdf.cell(0, 20, 'Power Measurement', 0, 1, 'C')
    pdf.set_font('Times', '', 12)
    tempConfigurationIndex = 0
    for d in rawData:
        if (tempConfigurationIndex == 0): 
            pdf.cell(60, 10, 'SE=OFF SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 1): 
            pdf.cell(60, 10, 'SE=ON SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 2): 
            pdf.cell(60, 10, 'SE=OFF SEDC=ON', 1, 0, 'C')
            pdf.cell(60, 10, '900mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 3): 
            pdf.cell(60, 10, 'SE=OFF SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 4): 
            pdf.cell(60, 10, 'SE=ON SEDC=OFF', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        if (tempConfigurationIndex == 5): 
            pdf.cell(60, 10, 'SE=OFF SEDC=ON', 1, 0, 'C')
            pdf.cell(60, 10, '200mV (BL)', 1, 0, 'C')
        tempConfigurationIndex += 1
        pdf.cell(60, 10, '', 1, 1, 'C')
        pdf.cell(60, 10, 'Power Rail', 1, 0, 'C')
        pdf.cell(60, 10, 'Voltages (V)', 1, 0, 'C')
        pdf.cell(60, 10, 'Current (mA)', 1, 1, 'C')
        kl = list(d.keys())
        tempPower = 0
        tempIndex = 0;
        for onekey in kl:
            #print (onekey, d[onekey])
            pdf.cell(60, 10, onekey, 1, 0, 'C')
            pdf.cell(60, 10, str(d[onekey][0]), 1, 0, 'C')
            pdf.cell(60, 10, str(d[onekey][1]), 1, 1, 'C')
            tempPower += d[onekey][2]
            tempIndex += 1
            if (tempIndex%3 == 0):
                pdf.cell(60, 10, 'Power (mW/Ch)', 1, 0, 'C')
                pdf.cell(60, 10, str(tempPower), 1, 0, 'C')
                pdf.cell(60, 10, '', 1, 1, 'C')
                tempPower = 0
        pdf.cell(0, 10, '', 0, 1, 'C')

    for i in range(8):
        tempPlot = GetCurrentPlot(i+1,0)
        tempPlot.savefig('./output/plots/FE{}_900mV.png'.format(i+1))
        plt.close()
        tempPlot = GetCurrentPlot(i+1,1)
        tempPlot.savefig('./output/plots/FE{}_200mV.png'.format(i+1))
        plt.close()
        pdf.image('./output/plots/FE{}_200mV.png'.format(i+1), x = None, y = None, w = 150, h = 100, type = 'png', link = '')
        pdf.image('./output/plots/FE{}_900mV.png'.format(i+1), x = None, y = None, w = 150, h = 100, type = 'png', link = '')

    pdf.output('tuto1.pdf', 'F')

    MakePDF(1)

